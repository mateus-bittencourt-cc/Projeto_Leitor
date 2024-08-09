import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def formatar_patrimonio(numero):
    """Formata o número do patrimônio no formato 000.000"""
    return f"{numero:06d}"[:3] + "." + f"{numero:06d}"[3:]


def buscar_e_pintar(caminho_arquivo, numero_procurado):
    workbook = load_workbook(caminho_arquivo)
    fill_green = PatternFill(
        start_color="00FF00", end_color="00FF00", fill_type="solid"
    )
    patrimonio_encontrado = False
    log_text = ""

    for sheet in workbook.sheetnames:
        current_sheet = workbook[sheet]

        for row in current_sheet.iter_rows():
            for cell in row:
                if cell.value == numero_procurado:
                    for cell_in_row in row:
                        cell_in_row.fill = fill_green
                    log_text = f"Patrimônio {formatar_patrimonio(numero_procurado)} encontrado na seção '{sheet}' e pintado em verde."
                    patrimonio_encontrado = True
                    break
            if patrimonio_encontrado:
                break

    if not patrimonio_encontrado:
        log_text = f"Patrimônio {formatar_patrimonio(numero_procurado)} não encontrado."

    workbook.save(caminho_arquivo)
    return log_text, patrimonio_encontrado


def on_entry_change(event):
    numero_procurado = entry_patrimonio.get()

    try:
        numero_procurado = int(numero_procurado)
        caminho_arquivo = (
            r"C:\Users\Usuario\Desktop\cobaia-junio.xlsx"  # Atualize este caminho
        )
        log_text, patrimonio_encontrado = buscar_e_pintar(
            caminho_arquivo, numero_procurado
        )
        insert_log(log_text, patrimonio_encontrado)
    except ValueError:
        messagebox.showerror("Erro", "Código inválido. Por favor, insira um número.")
    finally:
        entry_patrimonio.delete(0, tk.END)


def insert_log(text, encontrado):
    if encontrado:
        log_text_widget.insert(tk.END, text + "\n", "encontrado")
    else:
        log_text_widget.insert(tk.END, text + "\n", "nao_encontrado")
    log_text_widget.yview(tk.END)  # Rola automaticamente para o final do texto


# Configuração da interface gráfica
root = tk.Tk()
root.title("Busca de Patrimônio")

# Criando os widgets
label_patrimonio = tk.Label(root, text="Digite o patrimônio:")
label_patrimonio.pack(pady=10)

entry_patrimonio = tk.Entry(root)
entry_patrimonio.pack(pady=10)

# Bind do evento KeyRelease para acionar a busca assim que a tecla "Enter" for pressionada
entry_patrimonio.bind("<Return>", on_entry_change)

# Criando o log de histórico de patrimônios lidos
log_label = tk.Label(root, text="Histórico de Patrimônios Lidos:")
log_label.pack(pady=10)

# Frame para conter o Text e as barras de rolagem
log_frame = tk.Frame(root)
log_frame.pack(pady=10)

# Barra de rolagem vertical
scrollbar_y = tk.Scrollbar(log_frame, orient=tk.VERTICAL)
scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

# Text para exibir o log
log_text_widget = tk.Text(
    log_frame, width=80, height=15, wrap=tk.WORD, yscrollcommand=scrollbar_y.set
)
log_text_widget.pack(side=tk.LEFT)

# Configurando a barra de rolagem
scrollbar_y.config(command=log_text_widget.yview)

# Tags para colorir o texto
log_text_widget.tag_configure("encontrado", background="light green")
log_text_widget.tag_configure("nao_encontrado", background="light coral")

# Iniciando o loop da interface
root.mainloop()
