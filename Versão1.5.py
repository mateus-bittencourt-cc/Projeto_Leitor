import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os


def formatar_para_comparacao(numero):
    """Remove pontos e formata o número para comparação"""
    return numero.replace(".", "")


def buscar_e_pintar(caminho_arquivo, numero_procurado, aba_selecionada):
    workbook = load_workbook(caminho_arquivo)
    fill_green = PatternFill(
        start_color="00FF00", end_color="00FF00", fill_type="solid"
    )
    fill_yellow = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    patrimonio_encontrado = False
    log_text = ""

    numero_procurado = formatar_para_comparacao(numero_procurado)

    if aba_selecionada in workbook.sheetnames:
        sheet = workbook[aba_selecionada]

        # Buscar na aba selecionada
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    valor_cell = formatar_para_comparacao(str(cell.value))
                    if numero_procurado in valor_cell:
                        for cell_in_row in row:
                            cell_in_row.fill = fill_green
                        log_text = f"Patrimônio {numero_procurado} encontrado na aba '{aba_selecionada}' e pintado em verde."
                        patrimonio_encontrado = True
                        break
            if patrimonio_encontrado:
                break

        if not patrimonio_encontrado:
            # Buscar nas outras abas
            for sheet_name in workbook.sheetnames:
                if sheet_name != aba_selecionada:
                    sheet = workbook[sheet_name]

                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value:
                                valor_cell = formatar_para_comparacao(str(cell.value))
                                if numero_procurado in valor_cell:
                                    for cell_in_row in row:
                                        cell_in_row.fill = fill_yellow
                                    log_text = f"Patrimônio {numero_procurado} encontrado na aba '{sheet_name}' e pintado em amarelo."
                                    patrimonio_encontrado = True
                                    break
                        if patrimonio_encontrado:
                            break

    if not patrimonio_encontrado:
        log_text = f"Patrimônio {numero_procurado} não encontrado em nenhuma aba."

    if patrimonio_encontrado:
        workbook.save(caminho_arquivo)
        atualizar_planilha_na_interface(caminho_arquivo, aba_selecionada)

    return log_text, patrimonio_encontrado


def on_entry_change(event):
    numero_procurado = entry_patrimonio.get()
    aba_selecionada = aba_combobox.get()

    if len(numero_procurado) < 5:
        messagebox.showerror(
            "Erro", "O código de patrimônio deve ter pelo menos 5 caracteres."
        )
        entry_patrimonio.delete(0, tk.END)
        return

    try:
        numero_procurado = formatar_para_comparacao(numero_procurado)
        log_text, patrimonio_encontrado = buscar_e_pintar(
            caminho_arquivo, numero_procurado, aba_selecionada
        )
        insert_log(log_text, patrimonio_encontrado)
    except ValueError:
        messagebox.showerror("Erro", "Código inválido. Por favor, insira um número.")
    finally:
        entry_patrimonio.delete(0, tk.END)


def insert_log(text, encontrado):
    if encontrado:
        if "amarelo" in text:
            log_text_widget.insert(tk.END, text + "\n", "amarelo")
        else:
            log_text_widget.insert(tk.END, text + "\n", "encontrado")
    else:
        log_text_widget.insert(tk.END, text + "\n", "nao_encontrado")
    log_text_widget.yview(tk.END)  # Rola automaticamente para o final do texto


def atualizar_planilha_na_interface(caminho_arquivo, aba_selecionada):
    # Limpar a Treeview
    for row in tree.get_children():
        tree.delete(row)

    # Carregar o arquivo Excel e a aba selecionada
    workbook = load_workbook(caminho_arquivo)
    if aba_selecionada in workbook.sheetnames:
        sheet = workbook[aba_selecionada]
        headers = [cell.value for cell in sheet[1]]

        # Configurar as colunas da Treeview
        tree["columns"] = headers
        for header in headers:
            tree.heading(header, text=header)
            tree.column(header, width=100)

        # Adicionar dados à Treeview
        for row in sheet.iter_rows(min_row=2):
            values = []
            for cell in row:
                values.append(cell.value)
            tree.insert("", tk.END, values=values)

        # Atualizar a visualização da planilha
        tree.update_idletasks()
    else:
        log_text_widget.insert(
            tk.END,
            f"A aba '{aba_selecionada}' não encontrada no arquivo.\n",
            "nao_encontrado",
        )


def on_aba_change(event):
    aba_selecionada = aba_combobox.get()
    atualizar_planilha_na_interface(caminho_arquivo, aba_selecionada)


# Caminho do arquivo na área de trabalho
desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
caminho_arquivo = os.path.join(desktop_path, "relatorio_pat_ic.xlsx")

# Configuração da interface gráfica
root = tk.Tk()
root.title("Busca de Patrimônio")

# Criando os widgets
label_patrimonio = tk.Label(root, text="Digite o patrimônio:")
label_patrimonio.pack(pady=10)

entry_patrimonio = tk.Entry(root)
entry_patrimonio.pack(pady=10)

# Criando a ComboBox para selecionar a aba
label_aba = tk.Label(root, text="Selecione a aba:")
label_aba.pack(pady=5)

aba_combobox = ttk.Combobox(root)
aba_combobox.pack(pady=5)
aba_combobox.bind("<<ComboboxSelected>>", on_aba_change)


# Função para atualizar a ComboBox com as abas disponíveis
def atualizar_abas():
    workbook = load_workbook(caminho_arquivo)
    abas = workbook.sheetnames
    aba_combobox["values"] = abas
    if abas:
        aba_combobox.set(abas[0])  # Definir a primeira aba como padrão


atualizar_abas()

# Bind do evento KeyRelease para acionar a busca assim que a tecla "Enter" for pressionada
entry_patrimonio.bind("<Return>", on_entry_change)

# Criando o log de histórico de patrimônios lidos
log_label = tk.Label(root, text="Histórico de Patrimônios Lidos:")
log_label.pack(pady=10)

# Frame para conter o Text e as barras de rolagem
log_frame = tk.Frame(root)
log_frame.pack(pady=10)

# Barra de rolagem vertical para o log
scrollbar_y = tk.Scrollbar(log_frame, orient=tk.VERTICAL)
scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

# Text para exibir o log
log_text_widget = tk.Text(
    log_frame, width=80, height=15, wrap=tk.WORD, yscrollcommand=scrollbar_y.set
)
log_text_widget.pack(side=tk.LEFT)

# Configurando a barra de rolagem
scrollbar_y.config(command=log_text_widget.yview)

# Tags para colorir o texto
log_text_widget.tag_configure("encontrado", background="light green")
log_text_widget.tag_configure("amarelo", background="yellow")
log_text_widget.tag_configure("nao_encontrado", background="light coral")

# Criando o frame para exibir a planilha
planilha_frame = tk.Frame(root)
planilha_frame.pack(pady=10, fill=tk.BOTH, expand=True)

# Treeview para exibir a planilha
tree = ttk.Treeview(planilha_frame, show="headings")
tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

# Barra de rolagem vertical para a Treeview
scrollbar_y_planilha = tk.Scrollbar(
    planilha_frame, orient=tk.VERTICAL, command=tree.yview
)
scrollbar_y_planilha.pack(side=tk.RIGHT, fill=tk.Y)
tree.config(yscrollcommand=scrollbar_y_planilha.set)

# Atualizar a visualização da planilha ao iniciar
atualizar_planilha_na_interface(caminho_arquivo, aba_combobox.get())

# Iniciando o loop da interface
root.mainloop()
