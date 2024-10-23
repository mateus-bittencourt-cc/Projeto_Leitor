import tkinter as tk
from tkinter import messagebox, filedialog, ttk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# Função para remover pontos e formatar o número para comparação
def formatar_para_comparacao(numero):
    return numero.replace(".", "")

# Função que busca o patrimônio e aplica cores conforme a aba e seção selecionadas
def buscar_e_pintar_por_secao(caminho_arquivo, numero_procurado, aba_selecionada, secao_procurada):
    workbook = load_workbook(caminho_arquivo)
    fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    patrimonio_encontrado_na_secao = False
    patrimonio_encontrado_em_outra_secao = False
    log_text = ""

    if aba_selecionada in workbook.sheetnames:
        sheet = workbook[aba_selecionada]

        # Verifica se há a coluna 'SEÇÃO' na planilha
        colunas = {cell.value.lower(): idx for idx, cell in enumerate(next(sheet.iter_rows())) if isinstance(cell.value, str)}
        if 'seção' in colunas:
            coluna_secao = colunas['seção']
        else:
            log_text = f"Coluna 'SEÇÃO' não encontrada na aba '{aba_selecionada}'."
            return log_text, False

        # Buscar na aba selecionada e filtrar pela seção
        for row in sheet.iter_rows(min_row=2):  # Ignorando a linha de cabeçalho
            valor_secao = row[coluna_secao].value
            if valor_secao:
                for cell in row:
                    if cell.value:
                        valor_cell = formatar_para_comparacao(str(cell.value))
                        if numero_procurado in valor_cell:
                            # Se encontrado na seção correta
                            if secao_procurada.lower() in str(valor_secao).lower():
                                for cell_in_row in row:
                                    cell_in_row.fill = fill_green
                                log_text = f"Patrimônio {numero_procurado} encontrado na seção '{secao_procurada}' da aba '{aba_selecionada}' e pintado em verde."
                                patrimonio_encontrado_na_secao = True
                            else:
                                # Se encontrado em outra seção
                                for cell_in_row in row:
                                    cell_in_row.fill = fill_yellow
                                log_text = f"Patrimônio {numero_procurado} encontrado na seção '{valor_secao}' (diferente de '{secao_procurada}') da aba '{aba_selecionada}' e pintado em amarelo."
                                patrimonio_encontrado_em_outra_secao = True
                            break

            if patrimonio_encontrado_na_secao or patrimonio_encontrado_em_outra_secao:
                break

    if not patrimonio_encontrado_na_secao and not patrimonio_encontrado_em_outra_secao:
        log_text = f"Patrimônio {numero_procurado} não encontrado na seção '{secao_procurada}' em nenhuma aba."

    if patrimonio_encontrado_na_secao or patrimonio_encontrado_em_outra_secao:
        workbook.save(caminho_arquivo)
        atualizar_planilha_na_interface(caminho_arquivo, aba_selecionada)

    return log_text, patrimonio_encontrado_na_secao or patrimonio_encontrado_em_outra_secao

# Função para inserir logs no widget de log
def insert_log(text, encontrado):
    log_text_widget.config(state=tk.NORMAL)
    if encontrado:
        if "amarelo" in text:
            log_text_widget.insert(tk.END, text + "\n", "amarelo")
        elif "verde" in text:
            log_text_widget.insert(tk.END, text + "\n", "verde")
    else:
        log_text_widget.insert(tk.END, text + "\n", "nao_encontrado")
    log_text_widget.config(state=tk.DISABLED)
    log_text_widget.yview(tk.END)  # Rola automaticamente para o final do texto

# Função para atualizar a interface ao carregar uma planilha
def atualizar_planilha_na_interface(caminho_arquivo, aba_selecionada):
    workbook = load_workbook(caminho_arquivo)
    if aba_selecionada in workbook.sheetnames:
        sheet = workbook[aba_selecionada]
        atualizar_secoes(sheet)
    else:
        insert_log(f"A aba '{aba_selecionada}' não foi encontrada no arquivo.\n", encontrado=False)

# Função para atualizar a lista de seções disponíveis no combobox de seções
def atualizar_secoes(sheet):
    secao_set = set()
    colunas = {cell.value.lower(): idx for idx, cell in enumerate(next(sheet.iter_rows())) if isinstance(cell.value, str)}
    if 'seção' in colunas:
        coluna_secao = colunas['seção']

        # Armazenar a seção selecionada atualmente
        secao_selecionada_anterior = secao_combobox.get()

        for row in sheet.iter_rows(min_row=2):  # Ignorando a linha de cabeçalho
            secao = row[coluna_secao].value
            if secao:
                secao_set.add(secao)

        secao_combobox["values"] = list(secao_set)

        # Manter a seção anteriormente selecionada se ela ainda existir
        if secao_selecionada_anterior in secao_set:
            secao_combobox.set(secao_selecionada_anterior)
        elif secao_set:
            secao_combobox.set(list(secao_set)[0])  # Define a primeira seção como padrão
    else:
        secao_combobox["values"] = []
        secao_combobox.set("")

# Função para atualizar ao selecionar uma nova aba
def on_aba_change(event):
    aba_selecionada = aba_combobox.get()
    atualizar_planilha_na_interface(caminho_arquivo, aba_selecionada)

# Função para carregar a planilha e suas abas
def carregar_planilha():
    global caminho_arquivo
    caminho_arquivo = filedialog.askopenfilename(
        filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os Arquivos", "*.*")]
    )
    if caminho_arquivo:
        atualizar_abas()
        atualizar_planilha_na_interface(caminho_arquivo, aba_combobox.get())
    else:
        messagebox.showwarning("Aviso", "Nenhuma planilha carregada.")

# Função para atualizar a lista de abas no combobox
def atualizar_abas():
    workbook = load_workbook(caminho_arquivo)
    abas = workbook.sheetnames
    aba_combobox["values"] = abas
    if abas:
        aba_combobox.set(abas[0])  # Definir a primeira aba como padrão

# Função para exportar o log para um arquivo Excel e aplicar cores
def exportar_log():
    try:
        caminho_arquivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os Arquivos", "*.*")]
        )
        if caminho_arquivo:
            # Verifica se o diretório existe
            diretorio = os.path.dirname(caminho_arquivo)
            if not os.path.exists(diretorio):
                messagebox.showerror("Erro", "O diretório não existe.")
                return

            workbook = load_workbook(caminho_arquivo)
            sheet = workbook.active

            log_entries = log_text_widget.get(1.0, tk.END).strip().split("\n")

            # Listas para armazenar as entradas categorizadas
            verdes = []
            amarelos = []
            vermelhos = []

            for entry in log_entries:
                if "encontrado na seção" in entry:  # Verde
                    verdes.append(entry)
                elif "encontrado em seção diferente" in entry:  # Amarelo
                    amarelos.append(entry)
                elif "não encontrado" in entry:  # Vermelho
                    vermelhos.append(entry)

            # Escreve as entradas organizadas na planilha e aplica cores
            linha = 1
            for entry in verdes:
                sheet.append([entry])
                sheet.cell(row=linha, column=1).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
                linha += 1

            for entry in amarelos:
                sheet.append([entry])
                sheet.cell(row=linha, column=1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                linha += 1

            for entry in vermelhos:
                sheet.append([entry])
                sheet.cell(row=linha, column=1).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                linha += 1

            workbook.save(caminho_arquivo)
            messagebox.showinfo("Exportação", "Log exportado com sucesso.")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao exportar log: {e}")

# Função para limpar o log
def limpar_log():
    log_text_widget.config(state=tk.NORMAL)
    log_text_widget.delete(1.0, tk.END)
    log_text_widget.config(state=tk.DISABLED)

# Função que lida com a busca quando o usuário clica no botão
def buscar_patrimonios():
    patrimônios = entry_patrimonio.get("1.0", tk.END).strip().splitlines()
    aba_selecionada = aba_combobox.get()
    secao_procurada = secao_combobox.get()

    for numero_procurado in patrimônios:
        if len(numero_procurado) < 5:
            messagebox.showerror("Erro", f"O código de patrimônio '{numero_procurado}' deve ter pelo menos 5 caracteres.")
            continue

        try:
            numero_procurado = formatar_para_comparacao(numero_procurado)
            log_text, patrimonio_encontrado = buscar_e_pintar_por_secao(
                caminho_arquivo, numero_procurado, aba_selecionada, secao_procurada
            )
            insert_log(log_text, patrimonio_encontrado)
        except ValueError:
            messagebox.showerror("Erro", f"Código inválido: '{numero_procurado}'. Por favor, insira um número válido.")

# Configuração inicial da janela
root = tk.Tk()
root.title("Busca e Pintura de Patrimônios")

# Frame para os botões
buttons_frame = tk.Frame(root)
buttons_frame.pack(pady=10)

carregar_planilha_button = tk.Button(buttons_frame, text="Carregar Planilha", command=carregar_planilha)
carregar_planilha_button.pack(side=tk.LEFT, padx=5)

exportar_log_button = tk.Button(buttons_frame, text="Exportar Log", command=exportar_log)
exportar_log_button.pack(side=tk.LEFT, padx=5)

limpar_log_button = tk.Button(buttons_frame, text="Limpar Log", command=limpar_log)
limpar_log_button.pack(side=tk.LEFT, padx=5)

# Frame para a seleção de aba e seção
busca_frame = tk.Frame(root)
busca_frame.pack(pady=10, fill=tk.X)

label_aba = tk.Label(busca_frame, text="Aba:")
label_aba.pack(side=tk.LEFT, padx=5)
aba_combobox = ttk.Combobox(busca_frame, state="readonly")
aba_combobox.pack(side=tk.LEFT, padx=5)
aba_combobox.bind("<<ComboboxSelected>>", on_aba_change)

label_patrimonio = tk.Label(busca_frame, text="Códigos dos Patrimônios:")
label_patrimonio.pack(side=tk.LEFT, padx=5)
entry_patrimonio = tk.Text(busca_frame, height=10, width=20)
entry_patrimonio.pack(side=tk.LEFT, padx=5)

# Botão para buscar patrimônios
buscar_button = tk.Button(busca_frame, text="Buscar Patrimônios", command=buscar_patrimonios)
buscar_button.pack(side=tk.LEFT, padx=5)

label_secao = tk.Label(busca_frame, text="Seção:")
label_secao.pack(side=tk.LEFT, padx=5)
secao_combobox = ttk.Combobox(busca_frame, state="readonly")
secao_combobox.pack(side=tk.LEFT, padx=5)

# Widget de log
log_text_widget = tk.Text(root, height=15, state=tk.DISABLED)
log_text_widget.pack(pady=10, fill=tk.BOTH, expand=True)

# Tags para colorir o texto do log
log_text_widget.tag_config("verde", foreground="green")
log_text_widget.tag_config("amarelo", foreground="orange")
log_text_widget.tag_config("nao_encontrado", foreground="red")

root.mainloop()
