import tkinter as tk
from tkinter import messagebox, filedialog, ttk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import csv

# Função para remover pontos e formatar o número para comparação
def formatar_para_comparacao(numero):
    return numero.replace(".", "")

# Função que busca o patrimônio e aplica cores conforme a aba e seção selecionadas
def buscar_e_pintar_por_secao(caminho_arquivo, numero_procurado, aba_selecionada, secao_procurada):
    workbook = load_workbook(caminho_arquivo)
    fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    patrimonio_encontrado = False
    log_text = ""

    numero_procurado = formatar_para_comparacao(numero_procurado)

    if aba_selecionada in workbook.sheetnames:
        sheet = workbook[aba_selecionada]

        # Verifica se há a coluna 'SEÇÃO' na planilha (com diferentes variações de maiúsculas/minúsculas)
        colunas = {cell.value.lower(): idx for idx, cell in enumerate(next(sheet.iter_rows())) if isinstance(cell.value, str)}
        if 'seção' in colunas:
            coluna_secao = colunas['seção']
        else:
            log_text = f"Coluna 'SEÇÃO' não encontrada na aba '{aba_selecionada}'."
            return log_text, False

        # Buscar na aba selecionada e filtrar pela seção
        for row in sheet.iter_rows(min_row=2):  # Ignorando a linha de cabeçalho
            valor_secao = row[coluna_secao].value
            if valor_secao and secao_procurada.lower() in str(valor_secao).lower():
                for cell in row:
                    if cell.value:
                        valor_cell = formatar_para_comparacao(str(cell.value))
                        if numero_procurado in valor_cell:
                            for cell_in_row in row:
                                cell_in_row.fill = fill_green
                            log_text = f"Patrimônio {numero_procurado} encontrado na seção '{secao_procurada}' da aba '{aba_selecionada}' e pintado em verde."
                            patrimonio_encontrado = True
                            break
            if patrimonio_encontrado:
                break

    if not patrimonio_encontrado:
        log_text = f"Patrimônio {numero_procurado} não encontrado na seção '{secao_procurada}' em nenhuma aba."

    if patrimonio_encontrado:
        workbook.save(caminho_arquivo)
        atualizar_planilha_na_interface(caminho_arquivo, aba_selecionada)

    return log_text, patrimonio_encontrado

# Função para inserir logs no widget de log
def insert_log(text, encontrado):
    log_text_widget.config(state=tk.NORMAL)
    if encontrado:
        if "amarelo" in text:
            log_text_widget.insert(tk.END, text + "\n", "amarelo")
        elif "verde" in text:
            log_text_widget.insert(tk.END, text + "\n", "verde")
    else:
        log_text_widget.insert(tk.END, text + "\n", "nao_encontrado")
    log_text_widget.config(state=tk.DISABLED)
    log_text_widget.yview(tk.END)  # Rola automaticamente para o final do texto

# Função para atualizar a interface ao carregar uma planilha
def atualizar_planilha_na_interface(caminho_arquivo, aba_selecionada):
    workbook = load_workbook(caminho_arquivo)
    if aba_selecionada in workbook.sheetnames:
        sheet = workbook[aba_selecionada]
        # Aqui pode ser adicionado código para processar ou exibir a planilha conforme necessário
    else:
        insert_log(f"A aba '{aba_selecionada}' não foi encontrada no arquivo.\n", encontrado=False)

# Função para atualizar ao selecionar uma nova aba
def on_aba_change(event):
    aba_selecionada = aba_combobox.get()
    atualizar_planilha_na_interface(caminho_arquivo, aba_selecionada)

# Função para carregar a planilha e suas abas
def carregar_planilha():
    global caminho_arquivo
    caminho_arquivo = filedialog.askopenfilename(
        filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os Arquivos", "*.*")]
    )
    if caminho_arquivo:
        atualizar_abas()
        atualizar_planilha_na_interface(caminho_arquivo, aba_combobox.get())
    else:
        messagebox.showwarning("Aviso", "Nenhuma planilha carregada.")

# Função para atualizar a lista de abas no combobox
def atualizar_abas():
    workbook = load_workbook(caminho_arquivo)
    abas = workbook.sheetnames
    aba_combobox["values"] = abas
    if abas:
        aba_combobox.set(abas[0])  # Definir a primeira aba como padrão

# Função para exportar o log para CSV
def exportar_log():
    try:
        caminho_arquivo = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("Arquivos CSV", "*.csv"), ("Todos os Arquivos", "*.*")],
        )
        if caminho_arquivo:
            with open(caminho_arquivo, "w", newline="") as file:
                writer = csv.writer(file)
                log_entries = log_text_widget.get(1.0, tk.END).strip().split("\n")
                for entry in log_entries:
                    writer.writerow([entry])
            messagebox.showinfo("Exportação", "Log exportado com sucesso.")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao exportar log: {e}")

# Função para limpar o log
def limpar_log():
    log_text_widget.config(state=tk.NORMAL)
    log_text_widget.delete(1.0, tk.END)
    log_text_widget.config(state=tk.DISABLED)

# Função que lida com a busca quando o usuário pressiona Enter
def on_entry_change(event):
    numero_procurado = entry_patrimonio.get()
    aba_selecionada = aba_combobox.get()
    secao_procurada = entry_secao.get()

    if len(numero_procurado) < 5:
        messagebox.showerror("Erro", "O código de patrimônio deve ter pelo menos 5 caracteres.")
        entry_patrimonio.delete(0, tk.END)
        return

    try:
        numero_procurado = formatar_para_comparacao(numero_procurado)
        log_text, patrimonio_encontrado = buscar_e_pintar_por_secao(
            caminho_arquivo, numero_procurado, aba_selecionada, secao_procurada
        )
        insert_log(log_text, patrimonio_encontrado)
    except ValueError:
        messagebox.showerror("Erro", "Código inválido. Por favor, insira um número.")
    finally:
        entry_patrimonio.delete(0, tk.END)

# Configuração inicial da interface gráfica
root = tk.Tk()
root.title("Busca de Patrimônio por Seção")

# Frame para os botões
buttons_frame = tk.Frame(root)
buttons_frame.pack(pady=10, fill=tk.X)

carregar_planilha_button = tk.Button(buttons_frame, text="Carregar Planilha", command=carregar_planilha)
carregar_planilha_button.pack(side=tk.LEFT, padx=5)

exportar_log_button = tk.Button(buttons_frame, text="Exportar Log", command=exportar_log)
exportar_log_button.pack(side=tk.LEFT, padx=5)

limpar_log_button = tk.Button(buttons_frame, text="Limpar Log", command=limpar_log)
limpar_log_button.pack(side=tk.LEFT, padx=5)

# Frame para seleção de aba
aba_frame = tk.Frame(root)
aba_frame.pack(pady=10, fill=tk.X)

label_aba = tk.Label(aba_frame, text="Selecione a aba:")
label_aba.pack(side=tk.LEFT, padx=5)
aba_combobox = ttk.Combobox(aba_frame, state="readonly")
aba_combobox.pack(side=tk.LEFT, padx=5)
aba_combobox.bind("<<ComboboxSelected>>", on_aba_change)

# Frame para buscar patrimônio
busca_frame = tk.Frame(root)
busca_frame.pack(pady=10, fill=tk.X)

label_patrimonio = tk.Label(busca_frame, text="Código do Patrimônio:")
label_patrimonio.pack(side=tk.LEFT, padx=5)

entry_patrimonio = tk.Entry(busca_frame)
entry_patrimonio.pack(side=tk.LEFT, padx=5)
entry_patrimonio.bind("<Return>", on_entry_change)

# Frame para buscar por seção
secao_frame = tk.Frame(root)
secao_frame.pack(pady=10, fill=tk.X)

label_secao = tk.Label(secao_frame, text="Seção:")
label_secao.pack(side=tk.LEFT, padx=5)

entry_secao = tk.Entry(secao_frame)
entry_secao.pack(side=tk.LEFT, padx=5)

# Frame para o log
log_frame = tk.Frame(root)
log_frame.pack(pady=10, fill=tk.BOTH, expand=True)

log_text_widget = tk.Text(log_frame, state=tk.DISABLED, height=10, width=80)
log_text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

log_text_widget.tag_configure("verde", foreground="black", background="green")
log_text_widget.tag_configure("amarelo", foreground="black", background="yellow")
log_text_widget.tag_configure("nao_encontrado", foreground="black", background="red")

root.mainloop()
