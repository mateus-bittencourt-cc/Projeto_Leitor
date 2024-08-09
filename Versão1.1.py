from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Função para buscar e pintar a linha inteira em todas as abas
def buscar_e_pintar(caminho_arquivo, numero_procurado):
    # Carregar o arquivo Excel
    workbook = load_workbook(caminho_arquivo)

    # Definir a cor verde para preenchimento
    fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Variável para verificar se o patrimônio foi encontrado
    patrimonio_encontrado = False

    # Iterar sobre todas as planilhas no arquivo
    for sheet in workbook.sheetnames:
        current_sheet = workbook[sheet]

        # Percorrer todas as células da planilha
        for row in current_sheet.iter_rows():
            for cell in row:
                if cell.value == numero_procurado:
                    # Pintar toda a linha de verde
                    for cell_in_row in row:
                        cell_in_row.fill = fill_green
                    print(f"O patrimônio {numero_procurado} foi localizado na seção '{sheet}' e foi marcado em verde.")
                    patrimonio_encontrado = True
                    break  # Sai do loop interno quando o patrimônio é encontrado

            if patrimonio_encontrado:
                break  # Sai do loop externo se o patrimônio foi encontrado

    if not patrimonio_encontrado:
        print(f"Patrimônio {numero_procurado} não encontrado.")

    # Salvar as alterações
    workbook.save(caminho_arquivo)

# Caminho do arquivo Excel
caminho_arquivo = "junio.xlsx"

while True:
    # Captura a entrada do código de barras
    codigo_barras = input("Leia o patrimônio: ")

    # Converter para inteiro, se for um número
    try:
        numero_procurado = int(codigo_barras)
    except ValueError:
        print("Código inválido. Por favor, leia um patrimônio.")
        continue

    # Chamar a função para buscar e pintar
    buscar_e_pintar(caminho_arquivo, numero_procurado)
