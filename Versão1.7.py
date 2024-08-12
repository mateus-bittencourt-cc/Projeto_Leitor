import tkinter as tk
from tkinter import messagebox, filedialog, ttk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import csv


def formatar_para_comparacao(numero):
    """Remove pontos e formata o número para comparação"""
    return numero.replace(".", "")


def buscar_e_pintar(caminho_arquivo, numero_procurado, aba_selecionada):
    workbook = load_workbook(caminho_arquivo)
    fill_green = PatternFill(
        start_color="00FF00", end_color="00FF00", fill_type="solid"
    )
    fill_yellow = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    fill_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    patrimonio_encontrado = False
    log_text = ""

    numero_procurado = formatar_para_comparacao(numero_procurado)

    if aba_selecionada in workbook.sheetnames:
        sheet = workbook[aba_selecionada]

        # Buscar na aba selecionada
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    valor_cell = formatar_para_comparacao(str(cell.value))
                    if numero_procurado in valor_cell:
                        for cell_in_row in row:
                            cell_in_row.fill = fill_green
                        log_text = f"Patrimônio {numero_procurado} encontrado na aba '{aba_selecionada}' e pintado em verde."
                        patrimonio_encontrado = True
                        break
            if patrimonio_encontrado:
                break

        if not patrimonio_encontrado:
            # Buscar nas outras abas
            for sheet_name in workbook.sheetnames:
                if sheet_name != aba_selecionada:
                    sheet = workbook[sheet_name]

                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value:
                                valor_cell = formatar_para_comparacao(str(cell.value))
                                if numero_procurado in valor_cell:
                                    for cell_in_row in row:
                                        cell_in_row.fill = fill_yellow
                                    log_text = f"Patrimônio {numero_procurado} encontrado na aba '{sheet_name}' e pintado em amarelo."
                                    patrimonio_encontrado = True
                                    break
                        if patrimonio_encontrado:
                            break

    if not patrimonio_encontrado:
        log_text = f"Patrimônio {numero_procurado} não encontrado em nenhuma aba."

    if patrimonio_encontrado:
        workbook.save(caminho_arquivo)
        atualizar_planilha_na_interface(caminho_arquivo, aba_selecionada)

    return log_text, patrimonio_encontrado


def on_entry_change(event):
    numero_procurado = entry_patrimonio.get()
    aba_selecionada = aba_combobox.get()

    if len(numero_procurado) < 5:
        messagebox.showerror(
            "Erro", "O código de patrimônio deve ter pelo menos 5 caracteres."
        )
        entry_patrimonio.delete(0, tk.END)
        return

    try:
        numero_procurado = formatar_para_comparacao(numero_procurado)
        log_text, patrimonio_encontrado = buscar_e_pintar(
            caminho_arquivo, numero_procurado, aba_selecionada
        )
        insert_log(log_text, patrimonio_encontrado)
    except ValueError:
        messagebox.showerror("Erro", "Código inválido. Por favor, insira um número.")
    finally:
        entry_patrimonio.delete(0, tk.END)


def insert_log(text, encontrado):
    log_text_widget.config(state=tk.NORMAL)
    if encontrado:
        if "amarelo" in text:
            log_text_widget.insert(tk.END, text + "\n", "amarelo")
        elif "verde" in text:
            log_text_widget.insert(tk.END, text + "\n", "verde")
    else:
        log_text_widget.insert(tk.END, text + "\n", "nao_encontrado")
    log_text_widget.config(state=tk.DISABLED)
    log_text_widget.yview(tk.END)  # Rola automaticamente para o final do texto


def atualizar_planilha_na_interface(caminho_arquivo, aba_selecionada):
    # A função agora não utiliza a Treeview
    # Carregar o arquivo Excel e a aba selecionada
    workbook = load_workbook(caminho_arquivo)
    if aba_selecionada in workbook.sheetnames:
        sheet = workbook[aba_selecionada]
        # Aqui você pode adicionar lógica para processar ou exibir a planilha conforme necessário
    else:
        insert_log(
            f"A aba '{aba_selecionada}' não foi encontrada no arquivo.\n",
            encontrado=False,
        )


def on_aba_change(event):
    aba_selecionada = aba_combobox.get()
    atualizar_planilha_na_interface(caminho_arquivo, aba_selecionada)


def carregar_planilha():
    global caminho_arquivo
    caminho_arquivo = filedialog.askopenfilename(
        filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os Arquivos", "*.*")]
    )
    if caminho_arquivo:
        atualizar_abas()
        atualizar_planilha_na_interface(caminho_arquivo, aba_combobox.get())
    else:
        messagebox.showwarning("Aviso", "Nenhuma planilha carregada.")


def atualizar_abas():
    workbook = load_workbook(caminho_arquivo)
    abas = workbook.sheetnames
    aba_combobox["values"] = abas
    if abas:
        aba_combobox.set(abas[0])  # Definir a primeira aba como padrão


def exportar_log():
    try:
        caminho_arquivo = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("Arquivos CSV", "*.csv"), ("Todos os Arquivos", "*.*")],
        )
        if caminho_arquivo:
            with open(caminho_arquivo, "w", newline="") as file:
                writer = csv.writer(file)
                log_entries = log_text_widget.get(1.0, tk.END).strip().split("\n")
                for entry in log_entries:
                    writer.writerow([entry])
            messagebox.showinfo("Exportação", "Log exportado com sucesso.")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao exportar log: {e}")


def limpar_log():
    log_text_widget.config(state=tk.NORMAL)
    log_text_widget.delete(1.0, tk.END)
    log_text_widget.config(state=tk.DISABLED)


# Caminho inicial da planilha
caminho_arquivo = ""

# Configuração da interface gráfica
root = tk.Tk()
root.title("Busca de Patrimônio")

# Frame para os botões
buttons_frame = tk.Frame(root)
buttons_frame.pack(pady=10, fill=tk.X)

carregar_planilha_button = tk.Button(
    buttons_frame, text="Carregar Planilha", command=carregar_planilha
)
carregar_planilha_button.pack(side=tk.LEFT, padx=5)

exportar_log_button = tk.Button(
    buttons_frame, text="Exportar Log", command=exportar_log
)
exportar_log_button.pack(side=tk.LEFT, padx=5)

limpar_log_button = tk.Button(buttons_frame, text="Limpar Log", command=limpar_log)
limpar_log_button.pack(side=tk.LEFT, padx=5)

# Frame para seleção de aba
aba_frame = tk.Frame(root)
aba_frame.pack(pady=10, fill=tk.X)

label_aba = tk.Label(aba_frame, text="Selecione a aba:")
label_aba.pack(side=tk.LEFT, padx=5)
aba_combobox = ttk.Combobox(aba_frame, state="readonly")
aba_combobox.pack(side=tk.LEFT, padx=5)
aba_combobox.bind("<<ComboboxSelected>>", on_aba_change)

# Frame para buscar patrimônio
busca_frame = tk.Frame(root)
busca_frame.pack(pady=10, fill=tk.X)

label_patrimonio = tk.Label(busca_frame, text="Código do Patrimônio:")
label_patrimonio.pack(side=tk.LEFT, padx=5)

entry_patrimonio = tk.Entry(busca_frame)
entry_patrimonio.pack(side=tk.LEFT, padx=5)
entry_patrimonio.bind("<Return>", on_entry_change)

# Frame para exibir a planilha
planilha_frame = tk.Frame(root)
planilha_frame.pack(pady=10, fill=tk.BOTH, expand=True)

# Frame para o log
log_frame = tk.Frame(root)
log_frame.pack(pady=10, fill=tk.BOTH, expand=True)

log_text_widget = tk.Text(log_frame, height=10, wrap="word", bg="white", fg="black")
log_text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
log_text_widget.config(state=tk.DISABLED)

scroll_x = tk.Scrollbar(log_frame, orient="horizontal", command=log_text_widget.xview)
scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
log_text_widget.configure(xscrollcommand=scroll_x.set)

# Configurar tags de log
log_text_widget.tag_configure("verde", foreground="black", background="lightgreen")
log_text_widget.tag_configure("amarelo", foreground="black", background="yellow")
log_text_widget.tag_configure("vermelho", foreground="black", background="orange")
log_text_widget.tag_configure(
    "nao_encontrado", foreground="black", background="lightcoral"
)

root.mainloop()
