import tkinter as tk
from tkinter import messagebox, filedialog, ttk
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os

# Função para remover pontos e formatar o número para comparação
def formatar_para_comparacao(numero):
    return numero.replace(".", "")

# Função que busca o patrimônio e aplica cores conforme a aba e seção selecionadas
def buscar_e_pintar_por_secao(caminho_arquivo, numero_procurado, aba_selecionada, secao_procurada):
    workbook = load_workbook(caminho_arquivo)
    fill_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    patrimonio_encontrado_na_secao = False
    patrimonio_encontrado_em_outra_secao = False
    log_text = ""

    if aba_selecionada in workbook.sheetnames:
        sheet = workbook[aba_selecionada]

        # Verifica se há a coluna 'SEÇÃO' na planilha
        colunas = {cell.value.lower(): idx for idx, cell in enumerate(next(sheet.iter_rows())) if isinstance(cell.value, str)}
        if 'seção' in colunas:
            coluna_secao = colunas['seção']
        else:
            log_text = f"Coluna 'SEÇÃO' não encontrada na aba '{aba_selecionada}'."
            return log_text, False

        # Buscar na aba selecionada e filtrar pela seção
        for row in sheet.iter_rows(min_row=2):  # Ignorando a linha de cabeçalho
            valor_secao = row[coluna_secao].value
            if valor_secao:
                for cell in row:
                    if cell.value:
                        valor_cell = formatar_para_comparacao(str(cell.value))
                        if numero_procurado in valor_cell:
                            # Se encontrado na seção correta
                            if secao_procurada.lower() in str(valor_secao).lower():
                                for cell_in_row in row:
                                    cell_in_row.fill = fill_green
                                log_text = f"Patrimônio {numero_procurado} encontrado na seção '{secao_procurada}' da aba '{aba_selecionada}' e pintado em verde."
                                patrimonio_encontrado_na_secao = True
                            else:
                                # Se encontrado em outra seção
                                for cell_in_row in row:
                                    cell_in_row.fill = fill_yellow
                                log_text = f"Patrimônio {numero_procurado} encontrado na seção '{valor_secao}' (diferente de '{secao_procurada}') da aba '{aba_selecionada}' e pintado em amarelo."
                                patrimonio_encontrado_em_outra_secao = True
                            break

            if patrimonio_encontrado_na_secao or patrimonio_encontrado_em_outra_secao:
                break

    if not patrimonio_encontrado_na_secao and not patrimonio_encontrado_em_outra_secao:
        log_text = f"Patrimônio {numero_procurado} não encontrado na seção '{secao_procurada}' em nenhuma aba."

    if patrimonio_encontrado_na_secao or patrimonio_encontrado_em_outra_secao:
        workbook.save(caminho_arquivo)
        atualizar_planilha_na_interface(caminho_arquivo, aba_selecionada)

    return log_text, patrimonio_encontrado_na_secao or patrimonio_encontrado_em_outra_secao

# Função para inserir logs no widget de log
def insert_log(text, encontrado):
    log_text_widget.config(state=tk.NORMAL)
    if encontrado:
        if "amarelo" in text:
            log_text_widget.insert(tk.END, text + "\n", "amarelo")
        elif "verde" in text:
            log_text_widget.insert(tk.END, text + "\n", "verde")
    else:
        log_text_widget.insert(tk.END, text + "\n", "nao_encontrado")
    log_text_widget.config(state=tk.DISABLED)
    log_text_widget.yview(tk.END)  # Rola automaticamente para o final do texto

# Função para atualizar a interface ao carregar uma planilha
def atualizar_planilha_na_interface(caminho_arquivo, aba_selecionada):
    workbook = load_workbook(caminho_arquivo)
    if aba_selecionada in workbook.sheetnames:
        sheet = workbook[aba_selecionada]
        atualizar_secoes(sheet)
    else:
        insert_log(f"A aba '{aba_selecionada}' não foi encontrada no arquivo.\n", encontrado=False)

# Função para atualizar a lista de seções disponíveis no combobox de seções
def atualizar_secoes(sheet):
    secao_set = set()
    colunas = {cell.value.lower(): idx for idx, cell in enumerate(next(sheet.iter_rows())) if isinstance(cell.value, str)}
    if 'seção' in colunas:
        coluna_secao = colunas['seção']

        # Armazenar a seção selecionada atualmente
        secao_selecionada_anterior = secao_combobox.get()

        for row in sheet.iter_rows(min_row=2):  # Ignorando a linha de cabeçalho
            secao = row[coluna_secao].value
            if secao:
                secao_set.add(secao)

        secao_combobox["values"] = list(secao_set)

        # Manter a seção anteriormente selecionada se ela ainda existir
        if secao_selecionada_anterior in secao_set:
            secao_combobox.set(secao_selecionada_anterior)
        elif secao_set:
            secao_combobox.set(list(secao_set)[0])  # Define a primeira seção como padrão
    else:
        secao_combobox["values"] = []
        secao_combobox.set("")

# Função para atualizar ao selecionar uma nova aba
def on_aba_change(event):
    aba_selecionada = aba_combobox.get()
    atualizar_planilha_na_interface(caminho_arquivo, aba_selecionada)

# Função para carregar a planilha e suas abas
def carregar_planilha():
    global caminho_arquivo
    caminho_arquivo = filedialog.askopenfilename(
        filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os Arquivos", "*.*")]
    )
    if caminho_arquivo:
        atualizar_abas()
        atualizar_planilha_na_interface(caminho_arquivo, aba_combobox.get())
    else:
        messagebox.showwarning("Aviso", "Nenhuma planilha carregada.")

# Função para atualizar a lista de abas no combobox
def atualizar_abas():
    workbook = load_workbook(caminho_arquivo)
    abas = workbook.sheetnames
    aba_combobox["values"] = abas
    if abas:
        aba_combobox.set(abas[0])  # Definir a primeira aba como padrão

# Função para formatar o número do patrimônio para a exportação
def formatar_numero_exportacao(numero):
    # Pega apenas os últimos 6 dígitos
    numero_limpo = ''.join(filter(str.isdigit, numero))[-6:]
    if len(numero_limpo) == 6:
        return f"{numero_limpo[:3]}.{numero_limpo[3:]}"
    return numero # Retorna o número original se não tiver 6 dígitos

# Função para exportar o log para um arquivo Excel e aplicar cores em abas específicas
def exportar_log():
    try:
        caminho_arquivo_export = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os Arquivos", "*.*")]
        )
        if caminho_arquivo_export:
            # Criar um novo arquivo Excel
            workbook = Workbook()

            # Criar as abas
            sheet_encontrados = workbook.active
            sheet_encontrados.title = 'Encontrados'
            sheet_local_incorreto = workbook.create_sheet('Local Incorreto')
            sheet_nao_encontrados = workbook.create_sheet('Não Encontrados')

            # Listas para armazenar os patrimônios categorizados
            verdes = []
            amarelos = []
            vermelhos = []

            log_entries = log_text_widget.get(1.0, tk.END).strip().split("\n")
            for entry in log_entries:
                if not entry:
                    continue
                
                # Extrai o número do patrimônio da linha de log
                try:
                    # Divide a entrada por espaços e pega o segundo elemento, que é o número.
                    numero_patrimonio = entry.split(' ')[1]
                    patrimonio_formatado = formatar_numero_exportacao(numero_patrimonio)

                    if "pintado em verde" in entry:
                        verdes.append(patrimonio_formatado)
                    elif "pintado em amarelo" in entry:
                        amarelos.append(patrimonio_formatado)
                    elif "não encontrado" in entry:
                        vermelhos.append(patrimonio_formatado)
                except IndexError:
                    # Ignora linhas de log que não seguem o formato esperado
                    continue

            # Preencher as abas com os patrimônios formatados e aplicar cores
            for linha, patrimonio in enumerate(verdes, 1):
                cell = sheet_encontrados.cell(row=linha, column=1, value=patrimonio)
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

            for linha, patrimonio in enumerate(amarelos, 1):
                cell = sheet_local_incorreto.cell(row=linha, column=1, value=patrimonio)
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            for linha, patrimonio in enumerate(vermelhos, 1):
                cell = sheet_nao_encontrados.cell(row=linha, column=1, value=patrimonio)
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            workbook.save(caminho_arquivo_export)
            messagebox.showinfo("Exportação", "Log exportado com sucesso.")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao exportar log: {e}")

# Função para limpar o log
def limpar_log():
    log_text_widget.config(state=tk.NORMAL)
    log_text_widget.delete(1.0, tk.END)
    log_text_widget.config(state=tk.DISABLED)

# Função que lida com a busca quando o usuário clica no botão
def buscar_patrimonios():
    patrimônios = entry_patrimonio.get("1.0", tk.END).strip().splitlines()
    aba_selecionada = aba_combobox.get()
    secao_procurada = secao_combobox.get()

    for numero_procurado in patrimônios:
        if len(numero_procurado) < 5:
            messagebox.showerror("Erro", f"O código de patrimônio '{numero_procurado}' deve ter pelo menos 5 caracteres.")
            continue

        try:
            numero_procurado = formatar_para_comparacao(numero_procurado)
            log_text, patrimonio_encontrado = buscar_e_pintar_por_secao(
                caminho_arquivo, numero_procurado, aba_selecionada, secao_procurada
            )
            insert_log(log_text, patrimonio_encontrado)
        except ValueError:
            messagebox.showerror("Erro", f"Código inválido: '{numero_procurado}'. Por favor, insira um número válido.")

# Configuração inicial da janela
root = tk.Tk()
root.title("Busca e Pintura de Patrimônios")

# Frame para os botões
buttons_frame = tk.Frame(root)
buttons_frame.pack(pady=5)

carregar_planilha_button = tk.Button(buttons_frame, text="Carregar Planilha", command=carregar_planilha)
carregar_planilha_button.pack(side=tk.LEFT, padx=5)

exportar_log_button = tk.Button(buttons_frame, text="Exportar Log", command=exportar_log)
exportar_log_button.pack(side=tk.LEFT, padx=5)

limpar_log_button = tk.Button(buttons_frame, text="Limpar Log", command=limpar_log)
limpar_log_button.pack(side=tk.LEFT, padx=5)

# Frame para a seleção de aba e seção
busca_frame = tk.Frame(root)
busca_frame.pack(pady=10, fill=tk.X)

label_aba = tk.Label(busca_frame, text="Aba:")
label_aba.pack(side=tk.LEFT, padx=5)
aba_combobox = ttk.Combobox(busca_frame, state="readonly")
aba_combobox.pack(side=tk.LEFT, padx=5)
aba_combobox.bind("<<ComboboxSelected>>", on_aba_change)

label_patrimonio = tk.Label(busca_frame, text="Códigos dos Patrimônios:")
label_patrimonio.pack(side=tk.LEFT, padx=5)
entry_patrimonio = tk.Text(busca_frame, height=0.4, width=20)
entry_patrimonio.pack(side=tk.LEFT, padx=5)

# Botão para buscar patrimônios
buscar_button = tk.Button(busca_frame, text="Buscar Patrimônios", command=buscar_patrimonios)
buscar_button.pack(side=tk.LEFT, padx=5)

label_secao = tk.Label(busca_frame, text="Seção:")
label_secao.pack(side=tk.LEFT, padx=5)
secao_combobox = ttk.Combobox(busca_frame, state="readonly")
secao_combobox.pack(side=tk.LEFT, padx=5)

# Widget de log
log_text_widget = tk.Text(root, height=15, state=tk.DISABLED)
log_text_widget.pack(pady=10, fill=tk.BOTH, expand=True)

# Tags para colorir o texto do log
log_text_widget.tag_config("verde", foreground="green")
log_text_widget.tag_config("amarelo", foreground="orange")
log_text_widget.tag_config("nao_encontrado", foreground="red")

if __name__ == "__main__":
    root.mainloop()


import pytest
import tkinter as tk
from unittest.mock import MagicMock, patch
from openpyxl import Workbook
from junio5 import atualizar_planilha_na_interface

# Pytest fixture to create a Tkinter root window for the tests
@pytest.fixture(scope="module")
def root():
    """Create a root Tkinter window for tests that need it."""
    root = tk.Tk()
    yield root
    root.destroy()

# Pytest fixture to create a temporary Excel file for testing
@pytest.fixture
def temp_excel_file(tmp_path):
    """Create a temporary Excel file with some data for testing."""
    file_path = tmp_path / "test_sheet.xlsx"
    workbook = Workbook()
    
    # Sheet that exists and has a 'seção' column
    sheet_with_secao = workbook.active
    sheet_with_secao.title = "AbaComSecao"
    sheet_with_secao.append(["ID", "Patrimonio", "SEÇÃO"])
    sheet_with_secao.append([1, "123.456", "TI"])
    sheet_with_secao.append([2, "789.012", "RH"])

    # Sheet that exists but does not have a 'seção' column
    sheet_without_secao = workbook.create_sheet("AbaSemSecao")
    sheet_without_secao.append(["ID", "Item"])
    sheet_without_secao.append([1, "Cadeira"])

    workbook.save(file_path)
    return str(file_path)

# Test case for when the selected sheet exists
@patch('junio5.atualizar_secoes')
@patch('junio5.insert_log')
def test_atualizar_planilha_na_interface_sheet_exists(mock_insert_log, mock_atualizar_secoes, temp_excel_file, root):
    """
    Tests that 'atualizar_secoes' is called when the selected sheet exists.
    """
    # Arrange
    caminho_arquivo = temp_excel_file
    aba_selecionada = "AbaComSecao"

    # Act
    atualizar_planilha_na_interface(caminho_arquivo, aba_selecionada)

    # Assert
    mock_atualizar_secoes.assert_called_once()
    # Verify the sheet passed to the mock has the correct title
    called_sheet = mock_atualizar_secoes.call_args[0][0]
    assert called_sheet.title == aba_selecionada
    mock_insert_log.assert_not_called()

# Test case for when the selected sheet does not exist
@patch('junio5.atualizar_secoes')
@patch('junio5.insert_log')
def test_atualizar_planilha_na_interface_sheet_not_exists(mock_insert_log, mock_atualizar_secoes, temp_excel_file, root):
    """
    Tests that 'insert_log' is called with an error when the sheet does not exist.
    """
    # Arrange
    caminho_arquivo = temp_excel_file
    aba_selecionada = "AbaInexistente"

    # Act
    atualizar_planilha_na_interface(caminho_arquivo, aba_selecionada)

    # Assert
    mock_atualizar_secoes.assert_not_called()
    mock_insert_log.assert_called_once_with(
        f"A aba '{aba_selecionada}' não foi encontrada no arquivo.\n", 
        encontrado=False
    )

# Test case for when the file path is invalid or the file is corrupted
@patch('junio5.atualizar_secoes')
@patch('junio5.insert_log')
def test_atualizar_planilha_na_interface_invalid_file(mock_insert_log, mock_atualizar_secoes, root):
    """
    Tests the function's behavior with an invalid file path.
    """
    # Arrange
    caminho_arquivo = "caminho/invalido/nao_existe.xlsx"
    aba_selecionada = "QualquerAba"

    # Act & Assert
    with pytest.raises(FileNotFoundError):
        atualizar_planilha_na_interface(caminho_arquivo, aba_selecionada)
    
    mock_atualizar_secoes.assert_not_called()
    mock_insert_log.assert_not_called()
